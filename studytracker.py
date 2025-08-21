import tkinter as tk
from tkinter import simpledialog, messagebox, ttk
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import os

file_name = "study_log.xlsx"

# ---------------- Excel Setup ----------------
if not os.path.exists(file_name):
    wb = Workbook()

    # Logs
    ws1 = wb.active
    ws1.title = "Logs"
    ws1.append(["Date", "Start Time", "End Time", "Activity", "Duration"])

    # Summary (by day)
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Date", "Total Study Time (HH:MM:SS)", "Change vs Previous Day"])

    # Daily Target
    ws_d = wb.create_sheet("DailyTarget")
    ws_d.append(["Date", "Target Hours", "Earned Hours", "Progress %"])

    # Weekly Target
    ws_w = wb.create_sheet("WeeklyTarget")
    ws_w.append(["Week Start", "Target Hours", "Earned Hours", "Progress %"])

    wb.save(file_name)

wb = load_workbook(file_name)
logs_ws = wb["Logs"]
summary_ws = wb["Summary"]

# Create missing sheets if user had an older file
if "DailyTarget" not in wb.sheetnames:
    ws_d = wb.create_sheet("DailyTarget")
    ws_d.append(["Date", "Target Hours", "Earned Hours", "Progress %"])
    wb.save(file_name)
if "WeeklyTarget" not in wb.sheetnames:
    ws_w = wb.create_sheet("WeeklyTarget")
    ws_w.append(["Week Start", "Target Hours", "Earned Hours", "Progress %"])
    wb.save(file_name)

daily_ws = wb["DailyTarget"]
weekly_ws = wb["WeeklyTarget"]

start_time = None

# ---------------- Helpers ----------------
def to_td(timestr):
    """Convert 'HH:MM:SS' -> timedelta"""
    if not timestr:
        return timedelta()
    parts = str(timestr).split(":")
    if len(parts) == 3:
        h, m, s = parts
        # s might be float-like; we take int of floor seconds
        return timedelta(hours=int(h), minutes=int(m), seconds=int(float(s)))
    return timedelta()

def get_week_start(date_obj):
    """Return Monday date for the given date (as date)"""
    return (date_obj - timedelta(days=date_obj.weekday())).date()

def add_or_update_target(ws, key, target_hours):
    """Ensure a row exists for key (date or week_start) and set target."""
    found = False
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[0].value == key:
            ws.cell(i, 2).value = target_hours
            # Keep earned & % as is
            found = True
            break
    if not found:
        row_num = ws.max_row + 1
        ws.cell(row_num, 1).value = key
        ws.cell(row_num, 2).value = target_hours
        ws.cell(row_num, 3).value = 0
        ws.cell(row_num, 4).value = "0%"

def add_time(ws, key, duration_hours):
    """Increase earned hours for key and recompute % based on target."""
    found = False
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[0].value == key:
            prev_earned = row[2].value or 0
            target = row[1].value or 0
            new_earned = (prev_earned or 0) + duration_hours
            percent = min(100, round((new_earned / target) * 100, 1)) if target else 0
            ws.cell(i, 3).value = round(new_earned, 2)
            ws.cell(i, 4).value = f"{percent}%"
            found = True
            break
    if not found:
        row_num = ws.max_row + 1
        ws.cell(row_num, 1).value = key
        ws.cell(row_num, 2).value = 0
        ws.cell(row_num, 3).value = round(duration_hours, 2)
        ws.cell(row_num, 4).value = "0%"

# ---------------- Session Controls ----------------
def start_session():
    global start_time
    start_time = datetime.now()
    status_label.config(text=f"Started at {start_time.strftime('%H:%M:%S')}")

def stop_session():
    global start_time
    if not start_time:
        messagebox.showwarning("Error", "Start the session first!")
        return

    end_time = datetime.now()
    duration = end_time - start_time
    today_str = str(datetime.today().date())

    # Ask optional activity
    activity = simpledialog.askstring("Activity", "What did you study? (Optional)")
    if activity is None:
        activity = ""

    # Save to Logs sheet
    logs_ws.append([
        today_str,
        start_time.strftime("%H:%M:%S"),
        end_time.strftime("%H:%M:%S"),
        activity,
        str(duration).split(".")[0]
    ])

    # --- Update Summary (per day) ---
    found = False
    for i, row in enumerate(summary_ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[0].value == today_str:
            old_total = row[1].value or "0:00:00"
            old_td = to_td(old_total)
            new_total = old_td + duration
            summary_ws.cell(i, 1).value = today_str
            summary_ws.cell(i, 2).value = str(new_total).split(".")[0]

            # comparison with previous day
            if i > 2:
                prev_td = to_td(summary_ws.cell(i - 1, 2).value)
                diff = new_total - prev_td
                sign = '+' if diff.total_seconds() >= 0 else '-'
                summary_ws.cell(i, 3).value = f"{sign}{str(abs(diff)).split('.')[0]}"
            found = True
            break

    if not found:
        new_total = duration
        row_num = summary_ws.max_row + 1
        summary_ws.cell(row_num, 1).value = today_str
        summary_ws.cell(row_num, 2).value = str(new_total).split(".")[0]
        if row_num > 2:
            prev_td = to_td(summary_ws.cell(row_num - 1, 2).value)
            diff = new_total - prev_td
            sign = '+' if diff.total_seconds() >= 0 else '-'
            summary_ws.cell(row_num, 3).value = f"{sign}{str(abs(diff)).split('.')[0]}"

    # --- Update Daily Target ---
    hours = duration.total_seconds() / 3600.0
    add_time(daily_ws, today_str, hours)

    # --- Update Weekly Target ---
    week_start = str(get_week_start(datetime.today()))
    add_time(weekly_ws, week_start, hours)

    wb.save(file_name)

    status_label.config(text=f"Session saved! Duration: {str(duration).split('.')[0]}")
    start_time = None

    # Auto refresh summary window if open
    if summary_win and summary_win.winfo_exists():
        refresh_summary()
    # Refresh in-app dashboards
    refresh_targets()

# ---------------- Summary (table window) ----------------
def refresh_summary():
    """Refresh table contents"""
    for row in tree.get_children():
        tree.delete(row)
    for row in summary_ws.iter_rows(min_row=2, values_only=True):
        tree.insert("", "end", values=row)

def view_summary():
    global summary_win, tree
    summary_win = tk.Toplevel(root)
    summary_win.title("Study Summary")
    summary_win.geometry("500x260")

    tree = ttk.Treeview(summary_win, columns=("Date", "Total", "Change"), show="headings")
    tree.heading("Date", text="Date")
    tree.heading("Total", text="Total Study Time")
    tree.heading("Change", text="Vs Previous Day")

    tree.column("Date", width=120, anchor="center")
    tree.column("Total", width=160, anchor="center")
    tree.column("Change", width=160, anchor="center")

    tree.pack(expand=True, fill="both")
    refresh_summary()

# ---------------- Targets UI + Logic ----------------
def set_daily_target():
    target = simpledialog.askinteger("Daily Target", "Enter target hours for today:")
    if not target:
        return
    today_str = str(datetime.today().date())
    add_or_update_target(daily_ws, today_str, target)
    wb.save(file_name)
    refresh_targets()

def set_weekly_target():
    target = simpledialog.askinteger("Weekly Target", "Enter target hours for this week:")
    if not target:
        return
    week_start = str(get_week_start(datetime.today()))
    add_or_update_target(weekly_ws, week_start, target)
    wb.save(file_name)
    refresh_targets()

def refresh_targets():
    """Update the in-app dashboard for daily & weekly targets."""
    today_str = str(datetime.today().date())
    week_start = str(get_week_start(datetime.today()))

    # ----- Daily -----
    d_target, d_earned, d_percent = 0, 0, 0
    for row in daily_ws.iter_rows(min_row=2, values_only=True):
        if row[0] == today_str:
            d_target = row[1] or 0
            d_earned = row[2] or 0
            d_percent = min(100, round((d_earned / d_target) * 100, 1)) if d_target else 0
            break

    daily_date.config(text=f"Date: {today_str}")
    daily_target_label.config(text=f"Target: {d_target} hrs")
    daily_earned_label.config(text=f"Earned: {d_earned:.2f} hrs")
    daily_bar["value"] = d_percent
    daily_pct_label.config(text=f"{d_percent:.1f}%")

    # ----- Weekly -----
    w_target, w_earned, w_percent = 0, 0, 0
    for row in weekly_ws.iter_rows(min_row=2, values_only=True):
        if row[0] == week_start:
            w_target = row[1] or 0
            w_earned = row[2] or 0
            w_percent = min(100, round((w_earned / w_target) * 100, 1)) if w_target else 0
            break

    weekly_date.config(text=f"Week Start: {week_start}")
    weekly_target_label.config(text=f"Target: {w_target} hrs")
    weekly_earned_label.config(text=f"Earned: {w_earned:.2f} hrs")
    weekly_bar["value"] = w_percent
    weekly_pct_label.config(text=f"{w_percent:.1f}%")

# ---------------- GUI ----------------
root = tk.Tk()
root.title("Study Tracker")
root.geometry("420x640")

# Controls
start_btn = tk.Button(root, text="Start", command=start_session,
                      bg="green", fg="white", width=18, height=2)
start_btn.pack(pady=8)

stop_btn = tk.Button(root, text="Stop", command=stop_session,
                     bg="red", fg="white", width=18, height=2)
stop_btn.pack(pady=8)

summary_btn = tk.Button(root, text="View Summary", command=view_summary,
                        bg="blue", fg="white", width=18, height=2)
summary_btn.pack(pady=8)

daily_btn = tk.Button(root, text="Set Daily Target", command=set_daily_target,
                      bg="purple", fg="white", width=18, height=2)
daily_btn.pack(pady=6)

weekly_btn = tk.Button(root, text="Set Weekly Target", command=set_weekly_target,
                       bg="orange", fg="white", width=18, height=2)
weekly_btn.pack(pady=6)

status_label = tk.Label(root, text="Click Start to begin", fg="blue")
status_label.pack(pady=8)

# ---- Daily Progress Frame ----
daily_frame = tk.LabelFrame(root, text="Daily Progress", padx=8, pady=8)
daily_frame.pack(fill="x", padx=10, pady=6)

daily_date = tk.Label(daily_frame, text="Date: -")
daily_date.pack(anchor="w")
daily_target_label = tk.Label(daily_frame, text="Target: -")
daily_target_label.pack(anchor="w")
daily_earned_label = tk.Label(daily_frame, text="Earned: -")
daily_earned_label.pack(anchor="w")

daily_bar = ttk.Progressbar(daily_frame, orient="horizontal", length=360, mode="determinate", maximum=100)
daily_bar.pack(pady=6)
daily_pct_label = tk.Label(daily_frame, text="0.0%")
daily_pct_label.pack()

# ---- Weekly Progress Frame ----
weekly_frame = tk.LabelFrame(root, text="Weekly Progress", padx=8, pady=8)
weekly_frame.pack(fill="x", padx=10, pady=6)

weekly_date = tk.Label(weekly_frame, text="Week Start: -")
weekly_date.pack(anchor="w")
weekly_target_label = tk.Label(weekly_frame, text="Target: -")
weekly_target_label.pack(anchor="w")
weekly_earned_label = tk.Label(weekly_frame, text="Earned: -")
weekly_earned_label.pack(anchor="w")

weekly_bar = ttk.Progressbar(weekly_frame, orient="horizontal", length=360, mode="determinate", maximum=100)
weekly_bar.pack(pady=6)
weekly_pct_label = tk.Label(weekly_frame, text="0.0%")
weekly_pct_label.pack()

summary_win = None
tree = None

# Initialize dashboards
refresh_targets()

root.mainloop()
